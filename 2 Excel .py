from openpyxl import Workbook,load_workbook

excel1 = load_workbook("Excel1")
ws = excel1.active
excel2 = load_workbook("Excel2")
ws2 = excel2.active



mailler = []
for satir in range(1,ws.max_row+1):
    mailler.append(str(ws.cell(satir,1).value))

mailler2 = []
for satir2 in range(1,ws2.max_row+1):
    mailler2.append(str(ws2.cell(satir2,2).value))

adresler = []
for satir in range(1,ws2.max_row+1):
    adresler.append(str(ws2.cell(satir,3).value))

telefonNo = []
for satir in range(1,ws2.max_row+1):
    telefonNo.append(str(ws2.cell(satir,4).value))


abc = []
for i in mailler2:
    if i not in mailler:
        abc.append(i)
print(abc)
